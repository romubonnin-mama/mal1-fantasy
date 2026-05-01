import json
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

EXCEL_PATH = r"C:\Users\boro7\OneDrive\Documents\Draft Club\Saison 2025-2026.xlsm"

JOUEURS_CONFIG = {
    "ROMU":    {"col": 3,  "ligne": 4},
    "JEROME":  {"col": 3,  "ligne": 44},
    "VINCENT": {"col": 3,  "ligne": 84},
    "ADRIEN":  {"col": 22, "ligne": 4},
    "FLORIAN": {"col": 22, "ligne": 44},
    "FAB":     {"col": 22, "ligne": 84},
    "ANTHONY": {"col": 41, "ligne": 4},
    "BASTIEN": {"col": 41, "ligne": 44},
    "MICKA":   {"col": 41, "ligne": 84},
}

JOUEURS_CONFIG_ANCIEN = {
    "ROMU":    {"col": 3,  "ligne": 4},
    "JEROME":  {"col": 3,  "ligne": 44},
    "VINCENT": {"col": 3,  "ligne": 84},
    "ADRIEN":  {"col": 21, "ligne": 4},
    "FLORIAN": {"col": 21, "ligne": 44},
    "FAB":     {"col": 21, "ligne": 84},
    "ANTHONY": {"col": 39, "ligne": 4},
    "BASTIEN": {"col": 39, "ligne": 44},
    "MICKA":   {"col": 39, "ligne": 84},
}

# Baremes de scoring
BM_PTS   = {"G": 5, "D": 3, "M": 2, "A": 2}
PMA_PTS  = {"G": 2, "D": -2, "M": -2, "A": -2}
CS_PTS   = {"G": 2, "D": 1, "M": 0, "A": 0}
BE_PTS   = {"G": -2, "D": -1, "M": 0, "A": 0}


def to_int(v):
    return int(v) if isinstance(v, (int, float)) else 0


def calc_tj_pts(tj, red_card):
    s = str(tj).strip().upper() if tj else "0"
    if red_card:
        return 0
    if s == "M":
        return 4
    if s == "0" or not s:
        return 0
    if "-" in s:
        parts = s.split("-")
        try:
            minutes = int(parts[1]) - int(parts[0])
        except Exception:
            return 0
    else:
        try:
            minutes = int(s)
        except Exception:
            return 0
    if minutes <= 0:
        return 0
    elif minutes < 30:
        return 1
    elif minutes <= 60:
        return 2
    else:
        return 3


def calc_pts(poste, tj, bm_val, be_val, bcsc_val, cs_val, pm_val, pma_val, pd_val, cj_val, cr_val):
    red_card = (cr_val != 0)
    tj_pts   = calc_tj_pts(tj, red_card)
    bm_pts   = bm_val * BM_PTS.get(poste, 0)
    pd_pts   = pd_val * 2
    pm_pts   = pm_val * 2
    bcsc_pts = bcsc_val * (-2)
    cj_pts   = (0 if red_card else cj_val) * (-1)
    cr_pts   = 0
    pma_pts  = pma_val * PMA_PTS.get(poste, 0)
    cs_pts   = cs_val * CS_PTS.get(poste, 0)
    be_pts   = BE_PTS.get(poste, 0) if be_val >= 3 else 0
    total = tj_pts + bm_pts + pd_pts + pm_pts + pma_pts + cs_pts + be_pts + bcsc_pts + cj_pts + cr_pts
    return tj_pts, bm_pts, be_pts, bcsc_pts, cs_pts, pm_pts, pma_pts, pd_pts, cj_pts, cr_pts, total


def lire_classement(ws):
    joueurs = []
    for row in range(4, 13):
        rang = ws.cell(row=row, column=2).value
        nom  = ws.cell(row=row, column=3).value
        pts  = ws.cell(row=row, column=4).value
        if nom and isinstance(pts, (int, float)):
            joueurs.append({"rang": int(rang) if rang else row-3, "nom": str(nom), "pts": int(pts)})
    return sorted(joueurs, key=lambda x: x["rang"])


def lire_joueur(ws, col, row, ancien=False, poste="M"):
    decalage = 0 if ancien else 1
    nom = ws.cell(row=row, column=col).value
    if not nom or nom in ("", None):
        return None
    statut    = ws.cell(row=row, column=col+2).value
    cap       = ws.cell(row=row, column=col+3).value
    tj_entree = ws.cell(row=row, column=col+4).value
    tj_sortie = ws.cell(row=row, column=col+4+decalage).value
    if str(tj_sortie).upper() == "M":
        tj = "M"
    elif tj_entree and tj_sortie and tj_entree != tj_sortie:
        tj = f"{tj_entree}-{tj_sortie}"
    elif tj_sortie:
        tj = str(tj_sortie)
    else:
        tj = "0"

    bm_val   = to_int(ws.cell(row=row, column=col+5+decalage).value)
    be_val   = to_int(ws.cell(row=row, column=col+6+decalage).value)
    bcsc_val = to_int(ws.cell(row=row, column=col+7+decalage).value)
    cs_val   = to_int(ws.cell(row=row, column=col+8+decalage).value)
    pm_val   = to_int(ws.cell(row=row, column=col+9+decalage).value)
    pma_val  = to_int(ws.cell(row=row, column=col+10+decalage).value)
    pd_val   = to_int(ws.cell(row=row, column=col+11+decalage).value)
    cj_val   = to_int(ws.cell(row=row, column=col+12+decalage).value)
    cr_val   = to_int(ws.cell(row=row, column=col+13+decalage).value)

    tj_pts, bm_pts, be_pts, bcsc_pts, cs_pts, pm_pts, pma_pts, pd_pts, cj_pts, cr_pts, pts = \
        calc_pts(poste, tj, bm_val, be_val, bcsc_val, cs_val, pm_val, pma_val, pd_val, cj_val, cr_val)

    return {
        "nom":    str(nom),
        "statut": str(statut).lower() if statut else "",
        "cap":    str(cap) if cap else "",
        "tj":     tj,
        "tj_pts": tj_pts,
        "bm":     {"val": bm_val,   "pts": bm_pts},
        "be":     {"val": be_val,   "pts": be_pts},
        "bcsc":   {"val": bcsc_val, "pts": bcsc_pts},
        "cs":     {"val": cs_val,   "pts": cs_pts},
        "pm":     {"val": pm_val,   "pts": pm_pts},
        "pma":    {"val": pma_val,  "pts": pma_pts},
        "pd":     {"val": pd_val,   "pts": pd_pts},
        "cj":     {"val": cj_val,   "pts": cj_pts},
        "cr":     {"val": cr_val,   "pts": cr_pts},
        "pts":    pts,
    }


def lire_equipe(ws, nom_joueur, ancien=False):
    cfg   = (JOUEURS_CONFIG_ANCIEN if ancien else JOUEURS_CONFIG)[nom_joueur]
    col   = cfg["col"]
    ligne = cfg["ligne"]
    equipe = {"G": [], "D": [], "M": [], "A": []}

    g = lire_joueur(ws, col, ligne, ancien, poste="G")
    if g:
        equipe["G"].append(g)

    for r in range(ligne+3, ligne+15, 2):
        d = lire_joueur(ws, col, r, ancien, poste="D")
        if d:
            equipe["D"].append(d)

    for r in range(ligne+16, ligne+28, 2):
        m = lire_joueur(ws, col, r, ancien, poste="M")
        if m:
            equipe["M"].append(m)

    for r in range(ligne+29, ligne+37, 2):
        a = lire_joueur(ws, col, r, ancien, poste="A")
        if a:
            equipe["A"].append(a)

    return equipe


def main():
    print("Lecture du fichier Excel...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except FileNotFoundError:
        print("ERREUR : Fichier introuvable")
        input("Appuie sur Entree pour fermer...")
        return

    ws_scores  = wb["SCORES"]
    classement = lire_classement(ws_scores)

    historique      = {}
    detail_journees = {}

    for j in range(1, 35):
        if str(j) in wb.sheetnames and j != 38:
            ws_j   = wb[str(j)]
            ancien = (j < 14)
            detail_journees[str(j)] = {}
            for nom in JOUEURS_CONFIG:
                detail_journees[str(j)][nom] = lire_equipe(ws_j, nom, ancien)
            # Score journee = somme des pts de tous les joueurs
            historique[str(j)] = {
                nom: sum(p["pts"] for pos_players in equipe.values() for p in pos_players)
                for nom, equipe in detail_journees[str(j)].items()
            }
            print(f"  J{j} extraite")

    # Derniere journee = la plus haute avec au moins un score non nul
    derniere_j = max(
        (int(j) for j, scores in historique.items() if any(v > 0 for v in scores.values())),
        default=1
    )
    print(f"Derniere journee detectee : J{derniere_j}")

    joueurs_noms = [j["nom"] for j in classement]
    evolution    = {nom: [] for nom in joueurs_noms}
    cumul        = {nom: 0  for nom in joueurs_noms}
    for j in range(1, derniere_j + 1):
        if str(j) in historique:
            for nom in joueurs_noms:
                cumul[nom] += historique[str(j)].get(nom, 0)
            for nom in joueurs_noms:
                evolution[nom].append({"j": j, "pts": cumul[nom]})

    score_max_val = ws_scores.cell(row=14, column=9).value
    score_max_nom = ws_scores.cell(row=14, column=10).value
    score_min_val = ws_scores.cell(row=15, column=9).value
    score_min_nom = ws_scores.cell(row=15, column=10).value

    data = {
        "classement":      classement,
        "derniere_journee": derniere_j,
        "scores_journee":  historique[str(derniere_j)],
        "historique":      historique,
        "detail_journees": detail_journees,
        "evolution":       evolution,
        "score_max":       {"valeur": score_max_val, "joueur": score_max_nom},
        "score_min":       {"valeur": score_min_val, "joueur": score_min_nom},
    }

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print("Fichier data.json cree !")
    print("Envoi sur GitHub...")
    subprocess.run(["git", "add", "."], check=True)
    result = subprocess.run(["git", "commit", "-m", f"Mise a jour J{derniere_j}"])
    if result.returncode == 0:
        subprocess.run(["git", "push"], check=True)
    else:
        print("Aucun changement a envoyer.")
    print("Site mis a jour avec succes !")
    input("Appuie sur Entree pour fermer...")

if __name__ == "__main__":
    main()
