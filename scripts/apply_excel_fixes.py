"""
Ré-applique les corrections Excel après restauration d'une ancienne version :
  1. Formules CR anti-TJ (=IF(CRcol=1,-TJcol,0)) sur template 38 + J14-J32
  2. Code name VBA unique pour la feuille 32 (évite la suppression accidentelle de 38)
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
from export_excel import _save_preserving_images

EXCEL_PATH = Path(r"C:\Users\boro7\OneDrive\Documents\Draft Club\Saison 2025-2026.xlsm")

GROUP_HEADER_ROWS = [2, 42, 82]
POS_ROW_OFFSETS = {
    "G": [2],
    "D": [5, 7, 9, 11, 13],
    "M": [18, 20, 22, 24, 26, 28],
    "A": [31, 33, 35],
}


def update_cr_formulas(ws):
    count = skipped = 0
    for col_pos in range(3):
        name_col       = 2 + col_pos * 19
        tj_formula_col = name_col + 5
        cr_col         = name_col + 15
        cr_ltr = get_column_letter(cr_col)
        tj_ltr = get_column_letter(tj_formula_col)
        for hdr in GROUP_HEADER_ROWS:
            for offsets in POS_ROW_OFFSETS.values():
                for delta in offsets:
                    data_row    = hdr + delta
                    formula_row = data_row + 1
                    cell = ws.cell(row=formula_row, column=cr_col)
                    if isinstance(cell, MergedCell):
                        skipped += 1
                        continue
                    cell.value = f"=IF({cr_ltr}{data_row}=1,-{tj_ltr}{formula_row},0)"
                    count += 1
    return count, skipped


wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)

# 1. Formules CR sur template 38 et journées 14-37
for name in wb.sheetnames:
    if name == "38" or (name.isdigit() and 14 <= int(name) <= 37):
        n, s = update_cr_formulas(wb[name])
        print(f"  Sheet {name:>3}: {n} formules CR ok, {s} sautées (cellules fusionnées)")

# 2. Code name unique pour feuille 32
if "32" in wb.sheetnames:
    old_cn = wb["32"].sheet_properties.codeName
    wb["32"].sheet_properties.codeName = "J32"
    print(f"  Sheet 32 : codeName {old_cn!r} -> 'J32'")

_save_preserving_images(wb, EXCEL_PATH)
print("Sauvegarde OK - logos preserves, formules a jour.")
