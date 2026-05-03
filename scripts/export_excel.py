"""
Synchronise les stats admin (manual_stats.json + lineups.json) vers le fichier Excel.
Appelé automatiquement après chaque compute_journee.
"""

import copy
import io
import json
import posixpath
import re
import sys
import time
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
DATA_DIR = BASE_DIR / "data"
EXCEL_PATH = Path(r"C:\Users\boro7\OneDrive\Documents\Draft Club\Saison 2025-2026.xlsm")

# Position de chaque manager dans la grille (groupe, colonne)
# Groupe 0 = lignes 2-40, groupe 1 = lignes 42-80, groupe 2 = lignes 82-120
# Colonne 0=gauche, 1=milieu, 2=droite
MANAGER_GRID = {
    (0, 0): "ROMU",    (0, 1): "ADRIEN",  (0, 2): "ANTHONY",
    (1, 0): "JEROME",  (1, 1): "FLORIAN", (1, 2): "BASTIEN",
    (2, 0): "VINCENT", (2, 1): "FAB",     (2, 2): "MICKA",
}
MANAGER_TO_GRID = {v: k for k, v in MANAGER_GRID.items()}

# Ligne d'en-tête de chaque groupe
GROUP_HEADER_ROWS = [2, 42, 82]

# Offsets de ligne (depuis la ligne d'en-tête) pour chaque joueur
POS_ROW_OFFSETS = {
    "G": [2],
    "D": [5, 7, 9, 11, 13],
    "M": [18, 20, 22, 24, 26, 28],
    "A": [31, 33, 35],
}


def is_new_format(sheet_num: int) -> bool:
    return sheet_num >= 14


def get_name_col(col_pos: int, new_fmt: bool) -> int:
    """Colonne 1-based du label de position / nom manager."""
    block_width = 19 if new_fmt else 18
    return 2 + col_pos * block_width


def get_offsets(new_fmt: bool) -> dict:
    """Offsets depuis name_col vers chaque cellule d'entrée."""
    if new_fmt:
        return {
            "name": 1, "status": 3, "cap": 4, "entre": 5, "tj": 6,
            "bm": 7, "be": 8, "bcsc": 9, "cs": 10,
            "pm": 11, "pma": 12, "pd": 13, "cj": 14, "cr": 15,
        }
    else:
        return {
            "name": 1, "status": 2, "cap": 4, "tj": 5,
            "bm": 6, "be": 7, "bcsc": 8, "cs": 9,
            "pm": 10, "pma": 11, "pd": 12, "cj": 13, "cr": 14,
        }


def build_player_row_map(ws, group_idx: int, name_col: int, off: dict) -> dict:
    """Scanne la feuille pour trouver nom -> numéro de ligne."""
    header_row = GROUP_HEADER_ROWS[group_idx]
    col = name_col + off["name"]
    rows = {}
    for pos, offsets in POS_ROW_OFFSETS.items():
        for delta in offsets:
            row = header_row + delta
            val = ws.cell(row=row, column=col).value
            if val:
                rows[str(val).strip().upper()] = row
    return rows


def _minutes(s: dict) -> int:
    sort_a  = int(s.get("sort_a",  0) or 0)
    entre_a = int(s.get("entre_a", 0) or 0)
    fin_a   = int(s.get("fin_a",   0) or 0)
    if entre_a > 0:
        end = sort_a if sort_a > 0 else (fin_a if fin_a > 0 else 90)
        return end - entre_a
    if sort_a > 0:
        return sort_a
    return int(s.get("minutes", 0) or 0)


def compute_cs(poste: str, goals_conceded: int, minutes: int, full_match: bool) -> int:
    """Recalcule le clean sheet comme dans scoring.py."""
    if goals_conceded > 0:
        return 0
    if poste == "G" and (full_match or minutes >= 45):
        return 1
    if poste == "D" and (full_match or minutes > 45):
        return 1
    return 0


def _create_sheet(wb, journee: int, new_fmt: bool, off: dict):
    """Crée la feuille journée en copiant la feuille vierge '38' comme modèle."""
    if "38" not in wb.sheetnames:
        return None
    src = wb["38"]
    tgt = wb.copy_worksheet(src)
    tgt.title = str(journee)

    tgt.conditional_formatting = copy.deepcopy(src.conditional_formatting)

    # Zoom 72%
    tgt.sheet_view.zoomScale = 72

    # Code name VBA unique
    tgt.sheet_properties.codeName = f"J{journee}"

    # Déplacer la feuille à la bonne position (ordre numérique entre les journées)
    sheets = wb.sheetnames
    current_idx = sheets.index(str(journee))
    try:
        target_idx = next(i for i, s in enumerate(sheets) if s.isdigit() and int(s) > journee)
        wb.move_sheet(str(journee), offset=target_idx - current_idx)
    except StopIteration:
        idx_38 = sheets.index("38")
        wb.move_sheet(str(journee), offset=idx_38 - current_idx)
    return tgt


def _save_preserving_images(wb, path: Path, target_sheet: str = None) -> None:
    """
    Sauvegarde le workbook en préservant les images/drawings.

    Strategie : partir du ZIP original a 100%, patcher uniquement :
    - xl/workbook.xml         : ajout de la nouvelle feuille
    - xl/_rels/workbook.xml.rels : ajout du nouveau Relationship
    - [Content_Types].xml     : ajout Override pour la nouvelle feuille
    - target sheet XML        : injection du sheetData openpyxl
    Jamais de remplacement global par la version openpyxl (evite la corruption).
    """
    NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    NS_S = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

    path = Path(path)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if not path.exists():
        with open(path, 'wb') as f:
            f.write(buf.read())
        return

    def find_sheet_xml(zip_obj, sheet_name):
        """Retourne 'xl/worksheets/sheetN.xml' pour un nom de feuille."""
        try:
            root = ET.fromstring(zip_obj.read('xl/workbook.xml'))
            rid = None
            for el in root.iter(f'{{{NS_S}}}sheet'):
                if el.get('name') == sheet_name:
                    rid = el.get(f'{{{NS_R}}}id')
                    break
            if not rid:
                print(f"[img] '{sheet_name}' introuvable dans workbook.xml")
                return None
            rels_root = ET.fromstring(zip_obj.read('xl/_rels/workbook.xml.rels'))
            for rel in rels_root:
                if rel.get('Id') == rid:
                    target = rel.get('Target', '')
                    if target.startswith('/'):
                        return target.lstrip('/')
                    elif target.startswith('xl/'):
                        return target
                    else:
                        return posixpath.normpath('xl/' + target)
            return None
        except Exception as e:
            print(f"[img] find_sheet_xml('{sheet_name}') error: {e}")
            return None

    def extract_sheetdata(xml_text):
        m = re.search(r'(<sheetData(?:\s[^>]*)?>.*?</sheetData>)', xml_text, re.DOTALL)
        return m.group(1) if m else None

    def inject_sheetdata(base_xml, sheetdata):
        return re.sub(
            r'<sheetData(?:\s[^>]*)?>.*?</sheetData>',
            sheetdata, base_xml, count=1, flags=re.DOTALL)

    def add_missing_ns(base_xml, source_xml):
        src_ns = dict(re.findall(r'xmlns:(\w+)="([^"]+)"', source_xml[:3000]))
        base_head = base_xml[:3000]
        additions = ' '.join(
            f'xmlns:{pfx}="{uri}"'
            for pfx, uri in src_ns.items()
            if f'xmlns:{pfx}' not in base_head
        )
        if additions:
            base_xml = re.sub(
                r'(<worksheet\b[^>]*?)(/>|>)',
                lambda m: m.group(1) + ' ' + additions + m.group(2),
                base_xml, count=1)
        return base_xml

    result = io.BytesIO()
    with zipfile.ZipFile(path, 'r') as orig, \
         zipfile.ZipFile(buf, 'r') as new_z, \
         zipfile.ZipFile(result, 'w', zipfile.ZIP_DEFLATED) as out:

        orig_names = set(orig.namelist())

        # ── La feuille cible est-elle nouvelle (absente de l'original) ? ────────
        target_xml_orig = find_sheet_xml(orig, target_sheet) if target_sheet else None
        is_new_sheet = bool(target_sheet and target_xml_orig is None)
        print(f"[img] target='{target_sheet}', orig_xml={target_xml_orig}, is_new={is_new_sheet}")

        # ── SheetData de la feuille cible depuis la version openpyxl ────────────
        target_xml_new = find_sheet_xml(new_z, target_sheet) if target_sheet else None
        target_sheetdata = None
        target_new_xml_text = ''
        if target_xml_new and target_xml_new in new_z.namelist():
            target_new_xml_text = new_z.read(target_xml_new).decode('utf-8')
            target_sheetdata = extract_sheetdata(target_new_xml_text)
            print(f"[img] sheetData extrait de {target_xml_new} ({len(target_sheetdata or '')} chars)")

        # ── Preparation des fichiers a ajouter (nouvelle feuille seulement) ──────
        new_sheet_xml_name    = None
        new_sheet_xml_bytes   = None
        new_rels_path         = None
        new_sheet_rels_bytes  = None
        new_drawing_path      = None
        new_drawing_rels_path = None
        s38_drw_bytes         = None
        s38_drw_rels_bytes    = None
        patched_wb_xml        = None
        patched_wb_rels       = None
        patched_ct            = None

        if is_new_sheet:
            # Prochain numero de feuille XML disponible
            existing_nums = []
            for n in orig_names:
                mm = re.match(r'xl/worksheets/sheet(\d+)\.xml$', n)
                if mm:
                    existing_nums.append(int(mm.group(1)))
            next_num = max(existing_nums, default=0) + 1
            new_sheet_xml_name = f'xl/worksheets/sheet{next_num}.xml'
            new_rels_path = f'xl/worksheets/_rels/sheet{next_num}.xml.rels'
            print(f"[img] nouvelle feuille -> {new_sheet_xml_name}")

            # Drawing de la feuille modele '38'
            s38_xml = find_sheet_xml(orig, "38")
            s38_xml_text = None
            s38_drw_rid = 'rId1'
            if s38_xml and s38_xml in orig_names:
                s38_xml_text = orig.read(s38_xml).decode('utf-8')
                rels_38 = s38_xml.replace('xl/worksheets/', 'xl/worksheets/_rels/') + '.rels'
                if rels_38 in orig_names:
                    rels_root = ET.fromstring(orig.read(rels_38))
                    for rel in rels_root:
                        if 'drawing' in rel.get('Type', '').lower():
                            s38_drw_rid = rel.get('Id', 'rId1')
                            drw_target = rel.get('Target', '')
                            drw_path = posixpath.normpath(
                                posixpath.join('xl/worksheets', drw_target))
                            if drw_path in orig_names:
                                s38_drw_bytes = orig.read(drw_path)
                                drw_src_name = drw_path.split('/')[-1]
                                drw_rels_src = f'xl/drawings/_rels/{drw_src_name}.rels'
                                if drw_rels_src in orig_names:
                                    s38_drw_rels_bytes = orig.read(drw_rels_src)
                                    print(f"[img] drawing rels '38': {drw_rels_src} ({len(s38_drw_rels_bytes)} bytes)")
                                print(f"[img] drawing '38': {drw_path} ({len(s38_drw_bytes)} bytes)")
                            break
            print(f"[img] '38' XML: {s38_xml}, drawing: {s38_drw_bytes is not None}")

            # Chemin du nouveau drawing
            if s38_drw_bytes:
                all_drw = [n for n in orig_names
                           if re.match(r'xl/drawings/drawing\d+\.xml$', n)]
                max_drw = max(
                    (int(re.search(r'(\d+)', n).group(1)) for n in all_drw), default=0)
                new_drawing_path = f'xl/drawings/drawing{max_drw + 1}.xml'
                new_drw_fname = new_drawing_path.split('/')[-1]
                new_drawing_rels_path = f'xl/drawings/_rels/{new_drw_fname}.rels'
                print(f"[img] nouveau drawing: {new_drawing_path}")
                new_sheet_rels_bytes = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org'
                    '/package/2006/relationships">\n'
                    f'  <Relationship Id="{s38_drw_rid}"'
                    f' Type="http://schemas.openxmlformats.org/officeDocument'
                    f'/2006/relationships/drawing"'
                    f' Target="../drawings/{new_drw_fname}"/>\n'
                    '</Relationships>'
                ).encode('utf-8')

            # XML de la nouvelle feuille : base '38' + sheetData openpyxl
            if s38_xml_text and target_sheetdata:
                base = add_missing_ns(s38_xml_text, target_new_xml_text)
                base = inject_sheetdata(base, target_sheetdata)
                new_sheet_xml_bytes = base.encode('utf-8')
                print(f"[img] XML nouvelle feuille: '38' + sheetData patche")
            elif target_sheetdata:
                new_sheet_xml_bytes = target_new_xml_text.encode('utf-8')
                print(f"[img] XML nouvelle feuille: fallback openpyxl")

            # Nouveau rId pour workbook.xml.rels
            orig_rels_text = orig.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            existing_rids = [int(m) for m in re.findall(r'Id="rId(\d+)"', orig_rels_text)]
            next_rid_num = max(existing_rids, default=0) + 1
            new_rid = f'rId{next_rid_num}'
            print(f"[img] nouveau rId: {new_rid}")

            # Patch workbook.xml : ajout de l'entree <sheet>
            orig_wb_xml = orig.read('xl/workbook.xml').decode('utf-8')
            max_sid = max(
                (int(m) for m in re.findall(r'sheetId="(\d+)"', orig_wb_xml)), default=0)
            rel_target = 'worksheets/' + new_sheet_xml_name.split('/')[-1]
            new_sheet_entry = (
                f'<sheet name="{target_sheet}" sheetId="{max_sid + 1}"'
                f' r:id="{new_rid}"/>')
            # Insertion dans l'ordre numérique des onglets
            _inserted = False
            try:
                _num_j = int(target_sheet)
                for _m in re.finditer(r'<sheet\b[^>]*/>', orig_wb_xml):
                    _nm = re.search(r'\bname="([^"]*)"', _m.group(0))
                    if _nm:
                        try:
                            if int(_nm.group(1)) > _num_j:
                                _pos = _m.start()
                                patched_wb_xml = (orig_wb_xml[:_pos]
                                                  + new_sheet_entry + '\n    '
                                                  + orig_wb_xml[_pos:])
                                _inserted = True
                                break
                        except ValueError:
                            pass
            except ValueError:
                pass
            if not _inserted:
                patched_wb_xml = orig_wb_xml.replace(
                    '</sheets>', f'  {new_sheet_entry}\n</sheets>')
            print(f"[img] workbook.xml patche: sheetId={max_sid+1}, rId={new_rid}")

            # Patch workbook.xml.rels : ajout du Relationship
            new_rel_entry = (
                f'<Relationship Id="{new_rid}"'
                f' Type="http://schemas.openxmlformats.org/officeDocument'
                f'/2006/relationships/worksheet"'
                f' Target="{rel_target}"/>')
            patched_wb_rels = orig_rels_text.replace(
                '</Relationships>', f'  {new_rel_entry}\n</Relationships>')
            print(f"[img] workbook.xml.rels patche: {rel_target}")

            # Patch [Content_Types].xml : ajout Override feuille + drawing
            orig_ct = orig.read('[Content_Types].xml').decode('utf-8')
            sheet_ct = ('application/vnd.openxmlformats-officedocument'
                        '.spreadsheetml.worksheet+xml')
            new_part = '/' + new_sheet_xml_name
            if new_part not in orig_ct:
                orig_ct = orig_ct.replace(
                    '</Types>',
                    f'  <Override PartName="{new_part}" ContentType="{sheet_ct}"/>\n</Types>')
            if new_drawing_path:
                drw_ct = 'application/vnd.openxmlformats-officedocument.drawing+xml'
                drw_part = '/' + new_drawing_path
                if drw_part not in orig_ct:
                    orig_ct = orig_ct.replace(
                        '</Types>',
                        f'  <Override PartName="{drw_part}" ContentType="{drw_ct}"/>\n</Types>')
            patched_ct = orig_ct
            print(f"[img] [Content_Types].xml patche")

        # ── fullCalcOnLoad : force recalcul a l'ouverture (feuille nouvelle ET existante) ──
        _wb_base = patched_wb_xml if patched_wb_xml else orig.read('xl/workbook.xml').decode('utf-8')
        if 'fullCalcOnLoad="1"' not in _wb_base:
            _wb_patched = re.sub(
                r'(<calcPr\b[^>]*?)(/>|>)',
                lambda m: m.group(1) + ' fullCalcOnLoad="1"' + m.group(2),
                _wb_base, count=1)
            if _wb_patched != _wb_base:
                patched_wb_xml = _wb_patched
                print("[img] workbook.xml: fullCalcOnLoad=1 ajoute")

        # ── Ecriture du ZIP final : base = orig, quelques overrides ──────────────
        for name in orig.namelist():
            if name == 'xl/calcChain.xml':
                continue  # supprime : force Excel a recalculer toutes les formules a l'ouverture
            if name == 'xl/workbook.xml' and patched_wb_xml:
                out.writestr(name, patched_wb_xml.encode('utf-8'))
            elif name == 'xl/_rels/workbook.xml.rels' and patched_wb_rels:
                out.writestr(name, patched_wb_rels.encode('utf-8'))
            elif name == '[Content_Types].xml' and patched_ct:
                out.writestr(name, patched_ct.encode('utf-8'))
            elif (not is_new_sheet and target_xml_orig
                  and name == target_xml_orig and target_sheetdata):
                # Feuille cible existante : XML original + sheetData patche
                orig_xml = orig.read(name).decode('utf-8')
                if target_new_xml_text:
                    orig_xml = add_missing_ns(orig_xml, target_new_xml_text)
                out.writestr(name, inject_sheetdata(orig_xml, target_sheetdata).encode('utf-8'))
                print(f"[img] feuille cible '{target_sheet}': sheetData patche")
            else:
                out.writestr(name, orig.read(name))

        # ── Ajout des nouveaux fichiers (absents de l'original) ──────────────────
        if new_sheet_xml_name and new_sheet_xml_bytes:
            out.writestr(new_sheet_xml_name, new_sheet_xml_bytes)
            print(f"[img] ecrit: {new_sheet_xml_name}")

        if new_rels_path and new_sheet_rels_bytes:
            out.writestr(new_rels_path, new_sheet_rels_bytes)
            print(f"[img] ecrit: {new_rels_path}")

        if new_drawing_path and s38_drw_bytes:
            out.writestr(new_drawing_path, s38_drw_bytes)
            print(f"[img] ecrit: {new_drawing_path}")

        if new_drawing_rels_path and s38_drw_rels_bytes:
            out.writestr(new_drawing_rels_path, s38_drw_rels_bytes)
            print(f"[img] ecrit: {new_drawing_rels_path}")

    result.seek(0)
    data_to_write = result.read()
    for attempt in range(6):
        try:
            with open(path, 'wb') as f:
                f.write(data_to_write)
            break
        except PermissionError:
            if attempt < 5:
                print(f"[img] Fichier verrouille (OneDrive ?), retry dans 3s (tentative {attempt+1}/6)...")
                time.sleep(3)
            else:
                raise PermissionError(
                    f"Impossible d'ecrire '{path.name}' apres 6 tentatives. "
                    "Verifie qu'Excel et OneDrive ne bloquent pas le fichier."
                )


def export_journee(journee: int, verbose: bool = True) -> None:
    with open(DATA_DIR / "roster.json", encoding="utf-8") as f:
        roster = json.load(f)
    with open(DATA_DIR / "lineups.json", encoding="utf-8") as f:
        lineups = json.load(f)
    with open(DATA_DIR / "manual_stats.json", encoding="utf-8") as f:
        manual_stats = json.load(f)

    j_lineups = lineups.get(str(journee), {})
    j_stats = manual_stats.get(str(journee), {})

    if not j_lineups and not j_stats:
        if verbose:
            print(f"Aucune donnée admin pour la journée {journee}, export ignoré.")
        return

    sheet_name = str(journee)
    new_fmt = is_new_format(journee)
    off = get_offsets(new_fmt)

    wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)

    if sheet_name not in wb.sheetnames:
        ws = _create_sheet(wb, journee, new_fmt, off)
        if ws is None:
            if verbose:
                print(f"Impossible de créer la feuille '{sheet_name}'.")
            wb.close()
            return
        if verbose:
            print(f"Feuille '{sheet_name}' créée par copie.")

    ws = wb[sheet_name]

    for manager, player_positions in roster.items():
        grid_pos = MANAGER_TO_GRID.get(manager)
        if not grid_pos:
            continue

        group_idx, col_pos = grid_pos
        name_col = get_name_col(col_pos, new_fmt)
        player_rows = build_player_row_map(ws, group_idx, name_col, off)

        lineup     = j_lineups.get(manager, {})
        titulaires = set(lineup.get("titulaires", []))
        capitaine  = lineup.get("capitaine", "")
        coeff      = int(lineup.get("coeff", 1))
        m_stats    = j_stats.get(manager, {})

        for poste in ["G", "D", "M", "A"]:
            for player_idx, player in enumerate(player_positions.get(poste, [])):
                row = player_rows.get(player.strip().upper())
                if not row:
                    # Fallback pour feuille vierge (nouvelle) : position fixe
                    pos_offsets = POS_ROW_OFFSETS.get(poste, [])
                    if player_idx < len(pos_offsets):
                        row = GROUP_HEADER_ROWS[group_idx] + pos_offsets[player_idx]
                        ws.cell(row=row, column=name_col + off["name"]).value = player
                    else:
                        if verbose:
                            print(f"  [warn] {manager}/{player} introuvable dans feuille '{sheet_name}'")
                        continue

                is_titu = player in titulaires
                stats   = m_stats.get(player, {}) if is_titu else {}

                def w(field, value):
                    if field in off:
                        ws.cell(row=row, column=name_col + off[field]).value = value

                if not is_titu:
                    w("status", "r")
                    w("cap", None)
                    w("entre", None)
                    w("tj", None)
                    for field in ["bm","be","bcsc","cs","pm","pma","pd","cj","cr"]:
                        w(field, None)

                elif stats.get("absent"):
                    w("status", "A")
                    w("cap", None)
                    w("entre", None)
                    w("tj", None)
                    for field in ["bm","be","bcsc","cs","pm","pma","pd","cj","cr"]:
                        w(field, None)

                elif not stats:
                    pass

                else:
                    full_match = bool(stats.get("full_match", False))
                    entre_a    = int(stats.get("entre_a", 0) or 0)
                    sort_a     = int(stats.get("sort_a",  0) or 0)
                    fin_a      = int(stats.get("fin_a",   0) or 0)
                    red_card   = bool(stats.get("red_card", False))
                    minutes    = _minutes(stats)

                    if full_match:
                        w("entre", None)
                        w("tj",    "M")
                    elif entre_a > 0:
                        exit_min = sort_a if sort_a > 0 else (fin_a if fin_a > 0 else 90)
                        w("entre", entre_a)
                        w("tj",    exit_min)
                    elif sort_a > 0:
                        w("entre", None)
                        w("tj",    sort_a)
                    elif red_card and minutes > 0:
                        w("entre", None)
                        w("tj",    minutes)
                    else:
                        w("entre", None)
                        w("tj",    None)

                    w("status", None)
                    w("cap",    coeff if player == capitaine else None)

                    goals_c = int(stats.get("goals_conceded", 0))
                    cs_val  = 1 if stats.get("cs") else None
                    w("bm",   int(stats.get("goals", 0))        or None)
                    w("be",   goals_c                            or None)
                    w("bcsc", int(stats.get("own_goals", 0))     or None)
                    w("cs",   cs_val)
                    w("pm",   int(stats.get("pen_scored", 0))    or None)
                    w("pma",  int(stats.get("pen_mm_saved", 0))  or None)
                    w("pd",   int(stats.get("assists", 0))       or None)
                    w("cj",   int(stats.get("yellow_cards", 0))  or None)
                    w("cr",   1 if red_card                      else None)

    _save_preserving_images(wb, EXCEL_PATH, target_sheet=sheet_name)
    if verbose:
        print(f"Excel mis a jour : journee {journee} -> {EXCEL_PATH.name}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python export_excel.py <journee>")
        sys.exit(1)
    export_journee(int(sys.argv[1]))
