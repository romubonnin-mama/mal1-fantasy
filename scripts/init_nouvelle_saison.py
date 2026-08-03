"""
Initialise le fichier Excel pour la saison 2026-2027.

Actions :
  1. Copie Saison 2025-2026.xlsm -> Saison 2026-2027.xlsm
  2. Sur toutes les feuilles journee (1-38) et le template '38' :
     - Vide toutes les cellules de données joueurs (nom, statut, cap, TJ, stats)
     - Met à jour les entêtes manager dans la zone groupe 2 (ligne 82) :
       Vincent -> FAB  |  FAB -> MICKA  |  MICKA -> (vide)
  3. Remet les totaux SCORES à 0

Usage : python scripts/init_nouvelle_saison.py
"""

import shutil
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

SRC  = Path(r"C:\Users\boro7\OneDrive\Documents\Draft Club\Saison 2025-2026.xlsm")
DEST = Path(r"C:\Users\boro7\OneDrive\Documents\Draft Club\Saison 2026-2027.xlsm")

# Nouvelle grille managers 2026-2027
# Groupe 0 (ligne_header=2) : ROMU / ADRIEN / ANTHONY
# Groupe 1 (ligne_header=42): JEROME / FLORIAN / BASTIEN
# Groupe 2 (ligne_header=82): FAB / MICKA / (vide)
GROUP_HEADER_ROWS = [2, 42, 82]

# Offsets de ligne (depuis la ligne d'en-tête) pour chaque joueur
POS_ROW_OFFSETS = {
    "G": [2],
    "D": [5, 7, 9, 11, 13],
    "M": [18, 20, 22, 24, 26, 28],
    "A": [31, 33, 35],
}
ALL_OFFSETS = [o for offs in POS_ROW_OFFSETS.values() for o in offs]

# block_width et name_col (nouveau format j>=14)
NEW_BLOCK_WIDTH = 19
OLD_BLOCK_WIDTH = 18

# Offsets de colonnes depuis name_col (nouveau format)
NEW_OFF = {
    "name": 1, "status": 3, "cap": 4, "entre": 5, "tj": 6,
    "bm": 7, "be": 8, "bcsc": 9, "cs": 10,
    "pm": 11, "pma": 12, "pd": 13, "cj": 14, "cr": 15,
}
OLD_OFF = {
    "name": 1, "status": 2, "cap": 4, "tj": 5,
    "bm": 6, "be": 7, "bcsc": 8, "cs": 9,
    "pm": 10, "pma": 11, "pd": 12, "cj": 13, "cr": 14,
}
DATA_FIELDS = ["name", "status", "cap", "entre", "tj",
               "bm", "be", "bcsc", "cs", "pm", "pma", "pd", "cj", "cr"]

# Anciens noms managers dans le groupe 2 -> nouveaux noms
# col_pos 0 = col gauche (Vincent -> FAB), col_pos 1 = milieu (FAB -> MICKA), col_pos 2 = droite (MICKA -> vide)
RENAME_GROUP2 = {0: "FAB", 1: "MICKA", 2: None}


def get_name_col(col_pos: int, new_fmt: bool) -> int:
    bw = NEW_BLOCK_WIDTH if new_fmt else OLD_BLOCK_WIDTH
    return 2 + col_pos * bw


def clear_sheet(ws, new_fmt: bool):
    """Vide toutes les cellules de données joueurs dans une feuille."""
    off = NEW_OFF if new_fmt else OLD_OFF
    fields_to_clear = [f for f in DATA_FIELDS if f in off]

    for group_idx, header_row in enumerate(GROUP_HEADER_ROWS):
        for col_pos in range(3):
            name_col = get_name_col(col_pos, new_fmt)
            for delta in ALL_OFFSETS:
                row = header_row + delta
                for field in fields_to_clear:
                    ws.cell(row=row, column=name_col + off[field]).value = None


def update_manager_headers_group2(ws, new_fmt: bool):
    """
    Met à jour les cellules d'en-tête du groupe 2 (ligne 82) pour refléter
    la nouvelle disposition : FAB | MICKA | (vide).
    La cellule d'en-tête manager est supposée être à name_col + off['name'] - 1
    (la colonne 'pos/label' juste avant le nom).
    On cherche les cellules ligne 82 et 82+1 pour mettre à jour les titres.
    """
    header_row = GROUP_HEADER_ROWS[2]  # 82
    off = NEW_OFF if new_fmt else OLD_OFF

    for col_pos, new_name in RENAME_GROUP2.items():
        name_col = get_name_col(col_pos, new_fmt)
        # La cellule du nom manager est généralement dans la ligne d'en-tête
        # à la colonne name_col (label) ou name_col+1 (nom)
        # On met à jour la cellule titre (col = name_col)
        ws.cell(row=header_row, column=name_col).value = new_name


def reset_scores(ws):
    """Remet à 0 les totaux managers dans la feuille SCORES (col D, lignes 4-11)."""
    for row in range(4, 12):
        ws.cell(row=row, column=4).value = 0


def main():
    if not SRC.exists():
        print(f"ERREUR : fichier source introuvable : {SRC}")
        sys.exit(1)

    if DEST.exists():
        rep = input(f"'{DEST.name}' existe déjà. Écraser ? (o/N) : ").strip().lower()
        if rep != "o":
            print("Annulé.")
            sys.exit(0)

    print(f"Copie {SRC.name} -> {DEST.name} ...")
    shutil.copy2(SRC, DEST)
    print("Copie OK.")

    print("Ouverture du fichier Excel ...")
    wb = openpyxl.load_workbook(DEST, keep_vba=True)

    # Supprimer les feuilles journée (1-37) : elles seront créées automatiquement
    # par l'admin au fur et à mesure depuis le template '38'
    to_delete = [s for s in wb.sheetnames if s.isdigit() and s != "38"]
    for sheet_name in to_delete:
        del wb[sheet_name]
        print(f"  Feuille '{sheet_name}' supprimée.")

    # Nettoyer le template '38'
    if "38" in wb.sheetnames:
        clear_sheet(wb["38"], new_fmt=True)
        update_manager_headers_group2(wb["38"], new_fmt=True)
        print("  Template '38' nettoyé.")

    # Reset feuille SCORES
    if "SCORES" in wb.sheetnames:
        reset_scores(wb["SCORES"])
        print("  Feuille SCORES remise à 0.")

    print(f"Sauvegarde -> {DEST} ...")
    wb.save(DEST)
    print("Terminé. Fichier Excel 2026-2027 prêt.")
    print()
    print("Prochaines étapes :")
    print("  1. Ouvrir le fichier Excel et vérifier les onglets")
    print("  2. Saisir les effectifs après le draft (roster.json)")
    print("  3. Mettre à jour clubs.json après le draft")
    print("  4. Télécharger les nouveaux logos : python scripts/download_logos.py")


if __name__ == "__main__":
    main()
