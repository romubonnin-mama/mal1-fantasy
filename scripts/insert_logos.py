"""
Insere les logos des clubs dans le fichier Excel (toutes les feuilles).
Usage : python scripts/insert_logos.py
Prerequis : avoir lance download_logos.py d'abord.
"""

import json
import sys
import subprocess
from pathlib import Path

for pkg in ("openpyxl", "Pillow"):
    try:
        __import__(pkg.lower().replace("-", "_") if pkg != "Pillow" else "PIL")
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

BASE_DIR   = Path(__file__).parent.parent
LOGOS_DIR  = BASE_DIR / "logos"
EXCEL_PATH = Path(r"C:\Users\boro7\OneDrive\Documents\Draft Club\Saison 2025-2026.xlsm")

# Taille du logo en pixels (ajuster si trop grand / trop petit)
LOGO_PX = 22

with open(BASE_DIR / "clubs.json", encoding="utf-8") as f:
    clubs = {k.upper().replace("Ö", "O"): v for k, v in json.load(f).items()}

GROUP_HEADER_ROWS = [2, 42, 82]
POS_OFFSETS       = [2, 5, 7, 9, 11, 13, 18, 20, 22, 24, 26, 28, 31, 33, 35]


def name_col(col_pos: int, new_fmt: bool) -> int:
    return 2 + col_pos * (19 if new_fmt else 18)


def col_width_px(ws, letter: str) -> int:
    cd = ws.column_dimensions.get(letter)
    w = (cd.width if cd and cd.width else None) or ws.sheet_format.defaultColWidth or 8.43
    return max(int(w * 7), 5)


def merged_height_px(ws, row: int, col: int) -> int:
    letter = get_column_letter(col)
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= col <= mr.max_col and mr.min_row <= row <= mr.max_row:
            total = 0
            for r in range(mr.min_row, mr.max_row + 1):
                rd = ws.row_dimensions.get(r)
                h = (rd.height if rd and rd.height else None) or ws.sheet_format.defaultRowHeight or 15
                total += int(h * 96 / 72)
            return max(total, 5)
    rd = ws.row_dimensions.get(row)
    h = (rd.height if rd and rd.height else None) or ws.sheet_format.defaultRowHeight or 15
    return max(int(h * 96 / 72), 5)


if not EXCEL_PATH.exists():
    print(f"ERREUR : fichier introuvable : {EXCEL_PATH}")
    sys.exit(1)

print(f"Ouverture de {EXCEL_PATH.name} ...")
wb = openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)

for sheet_name in wb.sheetnames:
    if not sheet_name.isdigit():
        continue

    ws      = wb[sheet_name]
    n       = int(sheet_name)
    new_fmt = n >= 14

    # Supprimer les anciens logos
    ws._images.clear()

    count = 0
    for col_pos in range(3):
        base     = name_col(col_pos, new_fmt)
        ncol     = base + 1   # colonne nom joueur
        lcol     = base + 2   # colonne logo
        lletter  = get_column_letter(lcol)

        for hdr in GROUP_HEADER_ROWS:
            for off in POS_OFFSETS:
                row = hdr + off
                val = ws.cell(row=row, column=ncol).value
                if not val:
                    continue

                player   = str(val).strip().upper().replace("Ö", "O")
                club_id  = clubs.get(player)
                if not club_id:
                    continue

                logo_file = LOGOS_DIR / f"{club_id}.png"
                if not logo_file.exists():
                    continue

                # Taille : adapte a la cellule (ou fixe si trop petit)
                cell_w = col_width_px(ws, lletter)
                cell_h = merged_height_px(ws, row, lcol)
                marg   = 2
                lw     = max(min(LOGO_PX, cell_w  - marg * 2), 8)
                lh     = max(min(LOGO_PX, cell_h  - marg * 2), 8)

                img        = XLImage(str(logo_file))
                img.width  = lw
                img.height = lh
                img.anchor = f"{lletter}{row}"
                ws.add_image(img)
                count += 1

    print(f"  Feuille {sheet_name:>3} : {count} logos inseres")

print(f"\nSauvegarde ...")
wb.save(EXCEL_PATH)
print("Termine.")
