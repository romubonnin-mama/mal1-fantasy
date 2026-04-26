import json
import subprocess
import sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

EXCEL_PATH = r"C:\Users\boro7\OneDrive\Documents\Draft Club\Saison 2025-2026.xlsm"

def lire_classement(ws):
    joueurs = []
    for row in range(2, 11):
        rang  = ws.cell(row=row, column=2).value
        nom   = ws.cell(row=row, column=3).value
        pts   = ws.cell(row=row, column=4).value
        if nom and isinstance(pts, (int, float)):
            joueurs.append({"rang": int(rang) if rang else row-1, "nom": str(nom), "pts": int(pts)})
    return sorted(joueurs, key=lambda x: x["rang"])

def lire_scores_journee(ws, noms_cols):
    scores = {}
    for nom, col in noms_cols.items():
        # Le total est sur la ligne 2, dans la colonne TOTAL (col+17)
        val = ws.cell(row=2, column=col+17).value
        if isinstance(val, (int, float)):
            scores[nom] = int(val)
        else:
            scores[nom] = 0
    return scores

def derniere_journee_complete(wb):
    max_j = 0
    for sheet in wb.sheetnames:
        if sheet.isdigit():
            j = int(sheet)
            ws = wb[sheet]
            # Vérifie qu'au moins un score est renseigné (cellule TOTAL de ROMU = col 18, row 2)
            val = ws.cell(row=2, column=18).value
            if isinstance(val, (int, float)) and val > 0:
                max_j = max(max_j, j)
    return max_j if max_j > 0 else 1

def main():
    print("Lecture du fichier Excel...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except FileNotFoundError:
        print(f"ERREUR : Fichier introuvable :\n{EXCEL_PATH}")
        input("Appuie sur Entrée pour fermer...")
        return

    ws_scores = wb["SCORES"]
    classement = lire_classement(ws_scores)

    # Colonnes de départ pour chaque joueur (colonne 1 = A)
    noms_cols = {
        "ROMU":    1,
        "ADRIEN":  19,
        "ANTHONY": 37,
        "JEROME":  55,
        "FLORIAN": 73,
        "BASTIEN": 91,
        "VINCENT": 109,
        "FAB":     127,
        "MICKA":   145,
    }

    derniere_j = derniere_journee_complete(wb)
    print(f"Dernière journée détectée : J{derniere_j}")

    ws_j = wb[str(derniere_j)]
    scores_journee = lire_scores_journee(ws_j, noms_cols)

    # Score max et min depuis SCORES
    score_max_val  = ws_scores.cell(row=2, column=9).value
    score_max_nom  = ws_scores.cell(row=2, column=10).value
    score_min_val  = ws_scores.cell(row=3, column=9).value
    score_min_nom  = ws_scores.cell(row=3, column=10).value

    data = {
        "classement": classement,
        "derniere_journee": derniere_j,
        "scores_journee": scores_journee,
        "score_max": {"valeur": score_max_val, "joueur": score_max_nom},
        "score_min": {"valeur": score_min_val, "joueur": score_min_nom},
    }

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print("Fichier data.json créé !")
    print("Envoi sur GitHub...")
    subprocess.run(["git", "add", "."], check=True)
    subprocess.run(["git", "commit", "-m", f"Mise a jour J{derniere_j}"], check=True)
    subprocess.run(["git", "push"], check=True)
    print("Site mis à jour avec succès !")
    input("Appuie sur Entrée pour fermer...")

if __name__ == "__main__":
    main()